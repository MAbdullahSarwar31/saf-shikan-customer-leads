"""Tables component for SAF SHIKAN Lead Dashboard.

This module provides clean tabular displays and handles in-memory Excel generation.
"""

from io import BytesIO
import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def display_leads_table(df: pd.DataFrame) -> None:
    """Render a clean, sortable interactive table of lead data in Streamlit.
    
    Args:
        df: DataFrame containing lead columns.
    """
    # Select columns of interest for display
    display_cols = [
        "Priority_Rank", "Name", "Phone", "Crop_Type", "Crop_Area", 
        "Season", "Location", "Farm_Scale", "Lead_Score", 
        "Lead_Category", "Recommended_Action"
    ]
    
    # Ensure they exist in the input dataframe
    available_cols = [col for col in display_cols if col in df.columns]
    df_show = df[available_cols].copy()
    
    # Format crop types to title case for display
    if "Crop_Type" in df_show.columns:
        df_show["Crop_Type"] = df_show["Crop_Type"].str.title()
        
    st.dataframe(
        df_show,
        column_config={
            "Priority_Rank": st.column_config.NumberColumn(
                "Rank",
                help="Sales priority rank.",
                format="%d"
            ),
            "Name": "Farmer Name",
            "Phone": "Phone Number",
            "Crop_Type": "Crop Grown",
            "Crop_Area": st.column_config.NumberColumn(
                "Area (Acres)",
                help="Total crop area in acres.",
                format="%d"
            ),
            "Farm_Scale": "Farm Scale",
            "Lead_Score": st.column_config.ProgressColumn(
                "Lead Score",
                help="Lead Score out of 100",
                format="%.2f",
                min_value=0,
                max_value=100
            ),
            "Lead_Category": "Priority Category",
            "Recommended_Action": "Action Recommendation"
        },
        use_container_width=True,
        hide_index=True
    )


def generate_excel_bytes(df: pd.DataFrame) -> bytes:
    """Convert a DataFrame into a styled Excel workbook in memory and return its bytes.
    
    Args:
        df: Scored leads DataFrame.
        
    Returns:
        Excel file content in bytes.
    """
    export_cols = [
        "Priority_Rank", "Name", "Phone", "Crop_Type", "Crop_Area", 
        "Season", "Location", "Farm_Scale", "Lead_Score", 
        "Lead_Category", "Recommended_Action"
    ]
    # Filter to columns that exist
    available_cols = [col for col in export_cols if col in df.columns]
    export_df = df[available_cols].copy()
    
    if "Crop_Type" in export_df.columns:
        export_df["Crop_Type"] = export_df["Crop_Type"].str.title()
        
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Leads"
    ws.views.sheetView[0].showGridLines = True
    
    # Header Row
    headers = [col.replace("_", " ") for col in available_cols]
    ws.append(headers)
    
    # Styles
    header_fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid") # Dark Green
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    category_styles = {
        "HOT": {
            "fill": PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"),
            "font": Font(name="Calibri", size=10, bold=True, color="B71C1C")
        },
        "WARM": {
            "fill": PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid"),
            "font": Font(name="Calibri", size=10, bold=True, color="E65100")
        },
        "COLD": {
            "fill": PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"),
            "font": Font(name="Calibri", size=10, bold=True, color="0D47A1")
        },
        "LOW PRIORITY": {
            "fill": PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"),
            "font": Font(name="Calibri", size=10, color="757575")
        }
    }
    
    thin_border = Border(
        left=Side(style='thin', color='E0E0E0'),
        right=Side(style='thin', color='E0E0E0'),
        top=Side(style='thin', color='E0E0E0'),
        bottom=Side(style='thin', color='E0E0E0')
    )
    
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    
    # Style Header
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        
    # Write Rows
    for _, row in export_df.iterrows():
        ws.append(list(row))
        curr_row = ws.max_row
        category = row.get("Lead_Category", "")
        
        for col_idx in range(1, len(available_cols) + 1):
            cell = ws.cell(row=curr_row, column=col_idx)
            cell.border = thin_border
            cell.font = Font(name="Calibri", size=10, color="000000")
            
            col_name = available_cols[col_idx-1]
            if col_name in ["Priority_Rank", "Phone", "Crop_Area", "Season", "Farm_Scale", "Lead_Score", "Lead_Category"]:
                cell.alignment = center_align
            else:
                cell.alignment = left_align
                
            if col_name == "Lead_Score":
                cell.number_format = '0.00'
                
            if col_name == "Lead_Category" and category in category_styles:
                cell.fill = category_styles[category]["fill"]
                cell.font = category_styles[category]["font"]
                
    # Column width adjustments
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value or '')
            if cell.row > 1 and len(val) > max_len:
                max_len = len(val)
        
        header_len = len(str(col[0].value or ''))
        width = max(max_len + 3, header_len + 4)
        ws.column_dimensions[col_letter].width = min(width, 40)
        
    # Save to buffer
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_bytes = excel_buffer.getvalue()
    excel_buffer.close()
    
    return excel_bytes


def display_grouped_table(df: pd.DataFrame, group_col: str) -> None:
    """Group lead data and display styled aggregate statistics.
    
    Args:
        df: Customer Leads DataFrame.
        group_col: Column to group by.
    """
    import streamlit as st
    
    # Calculate grouped statistics
    agg_df = df.groupby(group_col).agg(
        Farmer_Count=('Name', 'count'),
        Avg_Lead_Score=('Lead_Score', 'mean'),
        Total_Crop_Area=('Crop_Area', 'sum'),
        Avg_Crop_Area=('Crop_Area', 'mean'),
        Avg_Income_PKR=('Estimated_Income', 'mean')
    ).reset_index()
    
    # Standardize crop naming capitalization
    group_col_title = group_col.replace("_", " ").title()
    if group_col == "Crop_Type":
        agg_df[group_col] = agg_df[group_col].str.title()
        
    # Rename columns for table display
    agg_df.columns = [
        group_col_title,
        "Farmers",
        "Average Lead Score",
        "Total Area (Acres)",
        "Average Area (Acres)",
        "Average Income (PKR)"
    ]
    
    # Format currency and score heatmap styling
    styled_df = agg_df.style.background_gradient(
        subset=["Average Lead Score"], 
        cmap="YlGn"
    ).format({
        "Average Lead Score": "{:.2f}",
        "Total Area (Acres)": "{:,.0f}",
        "Average Area (Acres)": "{:.1f}",
        "Average Income (PKR)": "PKR {:,.0f}"
    })
    
    st.dataframe(styled_df, use_container_width=True, hide_index=True)

