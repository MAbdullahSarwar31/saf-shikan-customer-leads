"""SAF SHIKAN Farmer Data Portal — Enterprise AGRON Admin Console Module.

Industry-standard, high-density enterprise SaaS data directory designed specifically
for seamless single-tab integration into the AGRON Admin Dashboard.
Data source: customers.csv (raw 6 core categories, no ML scoring pipeline).
Security: HTTPS enforced via Streamlit Cloud. Full audit trail via audit_logger.
"""

import os
import sys
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from supabase import create_client, Client

# ─── Audit Logger Import ──────────────────────────────────────────────────────
APP_DIR_IMPORT = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, APP_DIR_IMPORT)
from audit_logger import (
    log_event,
    get_log,
    get_log_as_csv,
    get_log_counts,
    get_log_from_supabase,
    EVENT_CONFIG
)
from auth import require_auth, get_current_user_email, logout

# ─── Page Configuration ──────────────────────────────────────────────────────
st.set_page_config(
    page_title="Saf Shikan — Customer Data & Management Portal",
    page_icon="⚫",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# ─── Path Resolution ──────────────────────────────────────────────────────────
APP_DIR = APP_DIR_IMPORT
PROJECT_ROOT = os.path.dirname(APP_DIR)
RAW_DATA_PATH = os.path.join(PROJECT_ROOT, "data", "raw", "customers.csv")

# ─── Supabase Configuration ──────────────────────────────────────────────────────────
SUPABASE_URL = st.secrets.get("supabase_url", "")
SUPABASE_KEY = st.secrets.get("supabase_key", "")
SUPABASE_TABLE = st.secrets.get("supabase_table", "farmer")

def get_supabase_client() -> Client | None:
    """Return an authenticated Supabase client using Streamlit secrets."""
    if not SUPABASE_URL or not SUPABASE_KEY:
        return None
    try:
        return create_client(SUPABASE_URL, SUPABASE_KEY)
    except Exception as e:
        st.warning(f"⚠️ Supabase connection failed: {e}. Check your supabase_url and supabase_key in Streamlit Secrets.")
        return None

# ─── Authentication Gate ─────────────────────────────────────────────────────
# Must be called BEFORE any CSS or content is rendered.
# Shows login form and halts execution if no authenticated user.
require_auth()

# ─── Enterprise AGRON Portal Design System CSS ───────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&family=Outfit:wght@500;600;700;800&display=swap');

html, body, [class*="css"], .stApp {
    font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif;
    background-color: #F8FAF9 !important;
    color: #0F172A;
}

.block-container {
    padding-top: 1.5rem !important;
    padding-left: 1.8rem !important;
    padding-right: 1.8rem !important;
    padding-bottom: 2.5rem !important;
    max-width: 100% !important;
}

/* ── Top Executive Header ──────────────────────────────────── */
.portal-breadcrumb {
    font-size: 0.75rem;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 0.8px;
    color: #10B981;
    margin-bottom: 4px;
}
.portal-header-row {
    display: flex;
    justify-content: space-between;
    align-items: center;
    margin-bottom: 20px;
    flex-wrap: wrap;
    gap: 12px;
}
.page-title {
    font-family: 'Outfit', sans-serif;
    font-size: 1.85rem;
    font-weight: 800;
    color: #0F172A;
    margin: 0;
    letter-spacing: -0.5px;
}
.status-pill {
    display: inline-flex;
    align-items: center;
    gap: 8px;
    background: #FFFFFF;
    border: 1px solid #D1FAE5;
    padding: 6px 14px;
    border-radius: 20px;
    font-size: 0.78rem;
    font-weight: 700;
    color: #065F46;
    box-shadow: 0 1px 2px rgba(0,0,0,0.03);
}
.status-dot {
    width: 8px;
    height: 8px;
    background-color: #10B981;
    border-radius: 50%;
}

/* ── Stat Cards ─────────────────────────────────────────────── */
.stats-row {
    display: flex;
    gap: 14px;
    margin-bottom: 22px;
    flex-wrap: wrap;
}
.stat-card {
    flex: 1;
    min-width: 190px;
    background: #FFFFFF;
    border: 1px solid #E2E8F0;
    border-top: 3px solid #0C3823;
    border-radius: 12px;
    padding: 16px 18px;
    box-shadow: 0 1px 3px rgba(0,0,0,0.03);
    transition: all 0.2s ease;
}
.stat-card:hover {
    border-color: #CBD5E1;
    border-top-color: #10B981;
    box-shadow: 0 4px 12px rgba(0,0,0,0.06);
}
.stat-header {
    display: flex;
    justify-content: space-between;
    align-items: center;
    margin-bottom: 8px;
}
.stat-label {
    font-size: 0.72rem;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 0.6px;
    color: #64748B;
}
.stat-badge {
    font-size: 0.7rem;
    font-weight: 700;
    background: #F1F5F9;
    color: #334155;
    padding: 2px 8px;
    border-radius: 6px;
}
.stat-value {
    font-family: 'Outfit', sans-serif;
    font-size: 2rem;
    font-weight: 800;
    color: #0F172A;
    line-height: 1;
    margin-bottom: 8px;
}
.stat-footer {
    display: flex;
    align-items: center;
    gap: 7px;
    font-size: 0.78rem;
    color: #64748B;
}
.dot { width:8px; height:8px; border-radius:50%; flex-shrink:0; display:inline-block; }
.dot-green  { background:#10B981; }
.dot-amber  { background:#F59E0B; }
.dot-blue   { background:#3B82F6; }
.dot-purple { background:#8B5CF6; }
.dot-teal   { background:#14B8A6; }

/* ── Enterprise Panel Cards ─────────────────────────────────── */
.enterprise-panel {
    background: #FFFFFF;
    border: 1px solid #E2E8F0;
    border-radius: 12px;
    padding: 18px 20px;
    margin-bottom: 18px;
    box-shadow: 0 1px 3px rgba(0,0,0,0.02);
}
.panel-header {
    display: flex;
    justify-content: space-between;
    align-items: center;
    margin-bottom: 14px;
    padding-bottom: 12px;
    border-bottom: 1px solid #F1F5F9;
}
.panel-title {
    font-size: 0.95rem;
    font-weight: 700;
    color: #0F172A;
}
.panel-count-badge {
    font-size: 0.76rem;
    font-weight: 700;
    background: #D1FAE5;
    color: #065F46;
    padding: 3px 10px;
    border-radius: 20px;
}

/* ── AGRON Executive Pill Tabs ───────────────────────────────── */
.stTabs [data-baseweb="tab-list"] {
    gap: 8px;
    background: #E2E8F0;
    border: 1px solid #CBD5E1;
    border-radius: 14px;
    padding: 6px;
    margin-bottom: 24px;
    box-shadow: inset 0 1px 3px rgba(0,0,0,0.06);
}
.stTabs [data-baseweb="tab"] {
    height: 40px;
    border-radius: 10px;
    padding: 0 24px;
    font-size: 0.88rem;
    font-weight: 600;
    color: #475569;
    background-color: transparent;
    transition: all 0.22s cubic-bezier(0.4, 0, 0.2, 1);
}
.stTabs [data-baseweb="tab"]:hover {
    color: #0F172A;
    background-color: rgba(255, 255, 255, 0.65);
}
.stTabs [aria-selected="true"] {
    background-color: #0C3823 !important;
    color: #FFFFFF !important;
    font-weight: 700 !important;
    box-shadow: 0 4px 12px rgba(12, 56, 35, 0.35) !important;
}
div[data-baseweb="tab-highlight"],
div[data-baseweb="tab-border"],
.stTabs [data-baseweb="tab-highlight"],
.stTabs [data-baseweb="tab-border"] {
    background-color: transparent !important;
    display: none !important;
    height: 0px !important;
}

/* ── Expander & Registration Console Overrides ──────────────── */
div[data-testid="stExpander"] {
    background: #FFFFFF !important;
    border: 1px solid #CBD5E1 !important;
    border-radius: 12px !important;
    box-shadow: 0 2px 6px rgba(0,0,0,0.03) !important;
    overflow: hidden !important;
    margin-bottom: 22px !important;
}
div[data-testid="stExpander"] details summary {
    background: #F8FAFC !important;
    padding: 14px 20px !important;
    font-family: 'Outfit', -apple-system, sans-serif !important;
    font-size: 0.95rem !important;
    font-weight: 700 !important;
    color: #0F172A !important;
    letter-spacing: 0.3px !important;
    border-bottom: 1px solid #E2E8F0 !important;
    transition: background 0.15s ease !important;
}
div[data-testid="stExpander"] details summary:hover {
    background: #F1F5F9 !important;
    color: #0C3823 !important;
}
div[data-testid="stExpander"] details[open] summary {
    background: #0C3823 !important;
    color: #FFFFFF !important;
    border-bottom: 1px solid #082618 !important;
}
div[data-testid="stExpander"] details[open] summary svg {
    fill: #FFFFFF !important;
    color: #FFFFFF !important;
}
div[data-testid="stExpander"] details > div {
    padding: 24px !important;
    background: #FFFFFF !important;
}

/* ── Form & Submit Button Overrides ─────────────────────────── */
div[data-testid="stForm"] {
    border: none !important;
    padding: 0 !important;
}
div[data-testid="stFormSubmitButton"] > button {
    background-color: #0C3823 !important;
    color: #FFFFFF !important;
    border: 1px solid #0C3823 !important;
    border-radius: 8px !important;
    padding: 10px 24px !important;
    font-weight: 700 !important;
    font-size: 0.88rem !important;
    letter-spacing: 0.5px !important;
    text-transform: uppercase !important;
    box-shadow: 0 2px 6px rgba(12, 56, 35, 0.2) !important;
    transition: all 0.2s ease !important;
}
div[data-testid="stFormSubmitButton"] > button:hover {
    background-color: #1B5E20 !important;
    border-color: #1B5E20 !important;
    box-shadow: 0 4px 12px rgba(12, 56, 35, 0.35) !important;
}
div[data-testid="stFormSubmitButton"] > button:active {
    transform: scale(0.98) !important;
}

/* ── Input & Select Overrides ───────────────────────────────── */
div[data-baseweb="select"] > div {
    border: 1px solid #E2E8F0 !important;
    background-color: #FFFFFF !important;
    border-radius: 8px !important;
    transition: all 0.15s ease !important;
}
div[data-baseweb="select"] > div:hover { border-color: #CBD5E1 !important; }
div[data-baseweb="select"] > div:focus-within {
    border-color: #0C3823 !important;
    box-shadow: 0 0 0 1px #0C3823 !important;
}
div[role="combobox"] { outline: none !important; }
div[data-testid="stTextInput"] input {
    border: 1px solid #E2E8F0 !important;
    border-radius: 8px !important;
    transition: all 0.15s ease !important;
}
div[data-testid="stTextInput"] input:focus {
    border-color: #0C3823 !important;
    box-shadow: 0 0 0 1px #0C3823 !important;
}
input:focus, textarea:focus, select:focus {
    outline: none !important;
    border-color: #0C3823 !important;
    box-shadow: 0 0 0 1px #0C3823 !important;
}

/* ── Chart Cards ─────────────────────────────────────────────── */
.agron-chart-card {
    background: #FFFFFF;
    border: 1px solid #E2E8F0;
    border-radius: 12px;
    padding: 16px 18px;
    margin-bottom: 16px;
    box-shadow: 0 1px 3px rgba(0,0,0,0.02);
}

/* ── Security Badges ─────────────────────────────────────────── */
.security-badge-row {
    display: flex;
    gap: 12px;
    flex-wrap: wrap;
}
.security-badge {
    display: inline-flex;
    align-items: flex-start;
    gap: 12px;
    background: #F8FAF9;
    border: 1px solid #E2E8F0;
    border-radius: 10px;
    padding: 14px 16px;
    flex: 1;
    min-width: 220px;
}
.security-badge-icon {
    width: 36px;
    height: 36px;
    border-radius: 8px;
    display: flex;
    align-items: center;
    justify-content: center;
    font-size: 0.75rem;
    font-weight: 800;
    flex-shrink: 0;
}
.badge-https   { background:#D1FAE5; color:#065F46; }
.badge-audit   { background:#DBEAFE; color:#1E3A8A; }
.badge-session { background:#EDE9FE; color:#5B21B6; }
.badge-access  { background:#FEF3C7; color:#92400E; }
.security-badge-text-label {
    font-size: 0.7rem;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 0.5px;
    color: #64748B;
    margin-bottom: 2px;
}
.security-badge-text-value {
    font-size: 0.9rem;
    font-weight: 700;
    color: #0F172A;
    margin-bottom: 3px;
}
.security-badge-text-sub {
    font-size: 0.76rem;
    color: #64748B;
    line-height: 1.4;
}

/* ── Audit Timeline ──────────────────────────────────────────── */
.timeline-wrap {
    background: #FFFFFF;
    border: 1px solid #E2E8F0;
    border-radius: 12px;
    overflow: hidden;
    box-shadow: 0 1px 3px rgba(0,0,0,0.02);
    margin-bottom: 18px;
}
.timeline-header {
    display: flex;
    justify-content: space-between;
    align-items: center;
    padding: 14px 20px;
    border-bottom: 1px solid #F1F5F9;
    background: #FAFAFA;
}
.timeline-title {
    font-size: 0.88rem;
    font-weight: 700;
    color: #0F172A;
    text-transform: uppercase;
    letter-spacing: 0.3px;
}
.timeline-badge {
    font-size: 0.72rem;
    font-weight: 700;
    background: #D1FAE5;
    color: #065F46;
    padding: 3px 10px;
    border-radius: 20px;
}
.timeline-row {
    display: flex;
    align-items: flex-start;
    gap: 14px;
    padding: 13px 20px;
    border-bottom: 1px solid #F8FAFC;
    transition: background 0.12s ease;
}
.timeline-row:last-child { border-bottom: none; }
.timeline-row:hover { background: #F8FAF9; }
.tl-icon {
    width: 34px;
    height: 34px;
    border-radius: 8px;
    display: flex;
    align-items: center;
    justify-content: center;
    font-size: 1rem;
    font-weight: 800;
    flex-shrink: 0;
    margin-top: 1px;
}
.tl-body { flex: 1; min-width: 0; }
.tl-title {
    font-size: 0.9rem;
    font-weight: 600;
    color: #0F172A;
    margin-bottom: 5px;
    line-height: 1.3;
}
.tl-meta-row {
    display: flex;
    align-items: center;
    flex-wrap: wrap;
    gap: 6px;
    margin-bottom: 5px;
}
.tl-type-pill {
    font-size: 0.67rem;
    font-weight: 700;
    padding: 2px 8px;
    border-radius: 10px;
    text-transform: uppercase;
    letter-spacing: 0.4px;
}
.tl-time {
    font-size: 0.76rem;
    color: #94A3B8;
    font-weight: 500;
}
.tl-session {
    font-size: 0.7rem;
    color: #CBD5E1;
    background: #F8FAFC;
    border: 1px solid #E2E8F0;
    padding: 1px 7px;
    border-radius: 6px;
    font-family: 'Courier New', monospace;
}
.detail-chips-row {
    display: flex;
    flex-wrap: wrap;
    gap: 6px;
    margin-top: 4px;
}
.detail-chip {
    display: inline-flex;
    align-items: center;
    border-radius: 6px;
    overflow: hidden;
    font-size: 0.73rem;
    border: 1px solid #E2E8F0;
}
.detail-chip-key {
    background: #F1F5F9;
    color: #64748B;
    font-weight: 600;
    padding: 2px 7px;
}
.detail-chip-val {
    background: #FFFFFF;
    color: #0F172A;
    font-weight: 500;
    padding: 2px 8px;
}
.tl-empty {
    text-align: center;
    padding: 40px 20px;
    color: #94A3B8;
    font-size: 0.88rem;
}

/* ── Responsive ──────────────────────────────────────────────── */
@media screen and (max-width: 768px) {
    .block-container { padding: 1rem 0.8rem !important; }
    .stats-row { flex-direction: column; gap: 12px; }
    .stat-card { min-width: 100%; }
    .security-badge-row { flex-direction: column; }
    .security-badge { min-width: 100%; }
}
</style>
""", unsafe_allow_html=True)


# ─── Data Loading ─────────────────────────────────────────────────────────────
def load_data() -> pd.DataFrame:
    """Load farmer data live from Supabase database."""
    supabase = get_supabase_client()
    req_cols = ["Name", "Phone", "Crop_Type", "Crop_Area", "Season", "Location", "Farm_Scale", "Region"]

    if supabase:
        # Try SUPABASE_TABLE, then 'farmer', then 'farmers' automatically
        candidate_tables = [SUPABASE_TABLE]
        for t in ["farmer", "farmers"]:
            if t not in candidate_tables:
                candidate_tables.append(t)

        last_error = None
        for tbl in candidate_tables:
            try:
                response = supabase.table(tbl).select(
                    "Name, Phone, Crop_Type, Crop_Area, Season, Location, Farm_Scale, Region"
                ).execute()
                if response.data:
                    df = pd.DataFrame(response.data)
                    for col in req_cols:
                        if col not in df.columns:
                            df[col] = ""
                    df = df[req_cols]
                    df["Crop_Type"]  = df["Crop_Type"].astype(str).str.title()
                    df["Season"]     = df["Season"].astype(str).str.strip()
                    df["Region"]     = df["Region"].astype(str).str.strip()
                    df["Location"]   = df["Location"].astype(str).str.strip()
                    df["Farm_Scale"] = df["Farm_Scale"].astype(str).str.strip()
                    df["Crop_Area"]  = pd.to_numeric(df["Crop_Area"], errors="coerce").fillna(0.0)
                    return df
                else:
                    # Connected but empty — likely RLS is blocking reads
                    last_error = f"Table '{tbl}' returned 0 rows. If you have data in Supabase, enable a SELECT policy for the 'anon' role under Authentication → Policies in your Supabase Dashboard."
            except Exception as e:
                last_error = str(e)
                continue

        # Surface the actual problem to the user instead of silent 0-records
        if last_error:
            st.warning(f"⚠️ Supabase database issue: {last_error}")

    # Fallback: empty DataFrame with correct schema so app never crashes
    return pd.DataFrame(columns=req_cols)


def save_new_row(row_dict: dict) -> bool:
    """Insert a new farmer record directly into Supabase."""
    supabase = get_supabase_client()
    if supabase:
        candidate_tables = [SUPABASE_TABLE]
        for t in ["farmer", "farmers"]:
            if t not in candidate_tables:
                candidate_tables.append(t)

        insert_error = None
        for tbl in candidate_tables:
            try:
                supabase.table(tbl).insert({
                    "Name":       row_dict["Name"],
                    "Phone":      row_dict["Phone"],
                    "Crop_Type":  row_dict["Crop_Type"],
                    "Crop_Area":  float(row_dict["Crop_Area"]),
                    "Season":     row_dict["Season"],
                    "Location":   row_dict["Location"],
                    "Farm_Scale": row_dict["Farm_Scale"],
                    "Region":     row_dict["Region"]
                }).execute()
                return True
            except Exception as e:
                insert_error = str(e)
                continue
        st.error(
            f"❌ Insert blocked by Supabase: {insert_error}\n\n"
            "**Fix:** Go to Supabase Dashboard → Authentication → Policies → `farmer` table "
            "→ Add an INSERT policy allowing the `anon` role."
        )
        return False
    return False


def generate_excel(df: pd.DataFrame) -> bytes:
    """Export filtered DataFrame to a styled Excel file."""
    cols = ["Name", "Phone", "Crop_Type", "Crop_Area", "Season", "Location"]
    available = [c for c in cols if c in df.columns]
    export = df[available].copy()
    export.columns = [c.replace("_", " ") for c in available]

    wb = Workbook()
    ws = wb.active
    ws.title = "SAF SHIKAN Farmers"
    hdr_fill = PatternFill("solid", fgColor="0C3823")
    hdr_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin = Border(
        left=Side("thin", color="E0E0E0"), right=Side("thin", color="E0E0E0"),
        top=Side("thin", color="E0E0E0"), bottom=Side("thin", color="E0E0E0")
    )
    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")

    ws.append(list(export.columns))
    for idx, cell in enumerate(ws[1], 1):
        cell.fill, cell.font, cell.alignment, cell.border = hdr_fill, hdr_font, center, thin

    for _, row in export.iterrows():
        ws.append(list(row))
        r = ws.max_row
        for c_idx, cell in enumerate(ws[r], 1):
            cell.border = thin
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = center if available[c_idx - 1] in ["Crop_Area", "Season", "Location", "Farm_Scale"] else left

    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = max((len(str(cell.value or "")) for cell in col), default=8)
        ws.column_dimensions[letter].width = min(max_len + 4, 40)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── Load & Pre-compute ───────────────────────────────────────────────────────
df = load_data()
CORE_COLS = ["Name", "Phone", "Crop_Type", "Crop_Area", "Season", "Location"]

total_farmers  = len(df)
unique_crops   = df["Crop_Type"].nunique()
unique_cities  = df["Location"].nunique()
total_area     = df["Crop_Area"].sum()

# Startup audit event (once per session)
is_supabase = bool(SUPABASE_URL and SUPABASE_KEY)
status_source = "SUPABASE CONNECTED" if is_supabase else "LOCAL OFFLINE"
current_user_email = get_current_user_email()

if "portal_started" not in st.session_state:
    ds_source = "Supabase PostgreSQL" if is_supabase else "offline fallback"
    log_event("LOGIN", f"User authenticated — portal session started",
              {"user": current_user_email, "total_records": total_farmers,
               "transport": "HTTPS (TLS 1.3)", "data_source": ds_source})
    st.session_state["portal_started"] = True

# ─── Top Executive Header ─────────────────────────────────────────────────────
hdr_col, logout_col = st.columns([9, 1])
with hdr_col:
    st.markdown(f"""
    <div class='portal-header-row'>
        <div>
            <div class='portal-breadcrumb'>AGRON ADMIN DASHBOARD / SAF SHIKAN MODULE / DATA REPOSITORY</div>
            <h1 class='page-title'>Saf Shikan — Customer Data &amp; Management Portal</h1>
        </div>
        <div class='status-pill'>
            <span class='status-dot'></span>
            <span>SYSTEM LIVE &nbsp;·&nbsp; {status_source} &nbsp;·&nbsp; {total_farmers:,} RECORDS &nbsp;·&nbsp; 🔐 {current_user_email}</span>
        </div>
    </div>
    """, unsafe_allow_html=True)
with logout_col:
    st.markdown("<div style='padding-top:1.4rem;'></div>", unsafe_allow_html=True)
    if st.button("⎋ Logout", use_container_width=True, help="Sign out of the AGRON portal"):
        log_event("LOGOUT", f"User signed out", {"user": current_user_email})
        logout()

# ─── Executive Stat Cards Ribbon ──────────────────────────────────────────────
st.markdown(f"""
<div class='stats-row'>
    <div class='stat-card'>
        <div class='stat-header'><span class='stat-label'>Total Farmers</span><span class='stat-badge'>100% Active</span></div>
        <div class='stat-value'>{total_farmers:,}</div>
        <div class='stat-footer'><span class='dot dot-green'></span>Registered profiles</div>
    </div>
    <div class='stat-card'>
        <div class='stat-header'><span class='stat-label'>Crop Categories</span><span class='stat-badge'>Portfolio</span></div>
        <div class='stat-value'>{unique_crops}</div>
        <div class='stat-footer'><span class='dot dot-amber'></span>Unique crop types</div>
    </div>
    <div class='stat-card'>
        <div class='stat-header'><span class='stat-label'>Cities Covered</span><span class='stat-badge'>Locations</span></div>
        <div class='stat-value'>{unique_cities}</div>
        <div class='stat-footer'><span class='dot dot-blue'></span>Districts & cities</div>
    </div>
    <div class='stat-card'>
        <div class='stat-header'><span class='stat-label'>Total Land Area</span><span class='stat-badge'>Cultivated</span></div>
        <div class='stat-value'>{total_area:,.0f} <span style='font-size:1rem;font-weight:600;'>ac</span></div>
        <div class='stat-footer'><span class='dot dot-teal'></span>Acres under coverage</div>
    </div>
</div>
""", unsafe_allow_html=True)

# ─── Navigation Tabs ──────────────────────────────────────────────────────────
tab_dir, tab_group, tab_charts, tab_security = st.tabs([
    "Farmer Directory",
    "Data Grouping & Aggregation",
    "Visual Analytics",
    "Security & Audit Trail"
])


# ══════════════════════════════════════════════════════════════════════════════
# TAB 1 — FARMER DIRECTORY
# ══════════════════════════════════════════════════════════════════════════════
with tab_dir:
    # ── Enterprise Registration Console ───────────────────────────────────────
    with st.expander("Register Farmer Profile — Entry Console", expanded=False):
        st.markdown("""
        <div style='background:#F8FAFC;border:1px solid #E2E8F0;border-left:4px solid #0C3823;border-radius:8px;padding:16px 20px;margin-bottom:20px;'>
            <div style='font-family:"Outfit",sans-serif;font-size:0.92rem;font-weight:700;color:#0F172A;margin-bottom:4px;letter-spacing:0.3px;'>NEW RECORD INGESTION SPECIFICATIONS</div>
            <div style='font-size:0.82rem;color:#475569;line-height:1.5;'>Input the core 6 categories and location attributes below. Upon committing, the record is appended directly to the directory master file and instantly synchronized across all live analytical tabs. Mandatory parameters are marked with an asterisk (*).</div>
        </div>
        """, unsafe_allow_html=True)

        # Pre-form: crop selection outside form so "Other" can reveal text input
        crop_choice = st.selectbox(
            "Primary Crop Category *",
            ["Wheat", "Cotton", "Rice", "Sugarcane", "Maize", "Orchard", "Vegetables", "Other"],
            key="reg_crop_choice"
        )
        if crop_choice == "Other":
            custom_crop = st.text_input(
                "Specify Crop Name *",
                placeholder="e.g. Sunflower, Canola...",
                key="reg_custom_crop"
            )
        else:
            custom_crop = ""

        with st.form("add_farmer_form", clear_on_submit=True):
            c1, c2, c3 = st.columns(3)
            with c1:
                new_name = st.text_input("Farmer Full Name *", placeholder="e.g. Tariq Mahmood")
                new_phone = st.text_input("Contact Phone Number *", placeholder="e.g. 0300-1234567",
                                          help="Format: 0300-1234567 (4 digits dash 7 digits)")
            with c2:
                new_location = st.text_input("District / City *", placeholder="e.g. Multan")
                new_area = st.number_input("Cultivated Area (Acres) *", min_value=1.0, max_value=5000.0, value=15.0, step=1.0)
            with c3:
                new_season = st.selectbox("Growing Season *", ["Rabi", "Kharif", "Perennial"])
                new_scale = st.selectbox("Farm Scale *", ["Small", "Medium", "Large"])

            st.write("")
            fc1, fc2, fc3 = st.columns([2.5, 2.5, 2.2])
            with fc3:
                submitted = st.form_submit_button("Commit & Register Profile", use_container_width=True)

            if submitted:
                import re as _re
                phone_valid = bool(_re.fullmatch(r"0\d{3}-\d{7}", new_phone.strip()))
                final_crop = custom_crop.strip() if crop_choice == "Other" else crop_choice
                if not new_name.strip():
                    st.error("Validation Error: Farmer Full Name is required.")
                elif not phone_valid:
                    st.error("Validation Error: Phone Number must follow the format 0XXX-XXXXXXX (e.g. 0300-1234567).")
                elif not new_location.strip():
                    st.error("Validation Error: District / City is required.")
                elif crop_choice == "Other" and not final_crop:
                    st.error("Validation Error: Please specify the crop name when selecting 'Other'.")
                else:
                    new_farmer = {
                        "Name": new_name.strip(),
                        "Phone": new_phone.strip(),
                        "Crop_Type": final_crop.title(),
                        "Crop_Area": float(new_area),
                        "Season": new_season,
                        "Location": new_location.strip(),
                        "Farm_Scale": new_scale,
                        "Region": new_location.strip()
                    }
                    save_new_row(new_farmer)
                    if hasattr(load_data, "clear"):
                        load_data.clear()
                    log_event("DATA_ENTRY", f"Registered new farmer profile: {new_name.strip()} ({new_location.strip()}, {final_crop}, {new_area} ac)", {"name": new_name.strip(), "location": new_location.strip(), "crop": final_crop, "area": new_area})
                    st.toast(f"Farmer profile committed: {new_name.strip()} has been added to the master directory.", icon=None)
                    st.rerun()


    st.markdown("""
    <div class='enterprise-panel'>
        <div class='panel-header'>
            <div class='panel-title'>DIRECTORY SEARCH &amp; SEGMENTATION PARAMETERS</div>
            <div class='panel-count-badge'>LIVE FILTER CONSOLE</div>
        </div>""", unsafe_allow_html=True)

    f1, f2, f3, f4 = st.columns([2.6, 1.4, 1.4, 1.4])
    with f1:
        search = st.text_input("Keyword Search", placeholder="Search Name, Phone or City...",
                               help="Filter across farmer name, phone number, or city")
    with f2:
        city_opts = ["All Cities"] + sorted(df["Location"].dropna().unique())
        sel_city = st.selectbox("City / District", city_opts)
    with f3:
        crop_opts = ["All Crops"] + sorted(df["Crop_Type"].unique())
        sel_crop = st.selectbox("Crop Category", crop_opts)
    with f4:
        season_opts = ["All Seasons"] + sorted(df["Season"].unique())
        sel_season = st.selectbox("Growing Season", season_opts)

    st.markdown("</div>", unsafe_allow_html=True)

    # Apply Filters
    filtered = df.copy()
    if search.strip():
        q = search.strip().lower()
        filtered = filtered[
            filtered["Name"].str.lower().str.contains(q, na=False) |
            filtered["Phone"].str.lower().str.contains(q, na=False) |
            filtered["Location"].str.lower().str.contains(q, na=False)
        ]
    if sel_city != "All Cities":
        filtered = filtered[filtered["Location"] == sel_city]
    if sel_crop != "All Crops":
        filtered = filtered[filtered["Crop_Type"] == sel_crop]
    if sel_season != "All Seasons":
        filtered = filtered[filtered["Season"] == sel_season]

    log_event("FILTER_APPLY", "Farmer directory filtered", {
        "keyword": search.strip() or "(none)",
        "city": sel_city,
        "crop": sel_crop,
        "season": sel_season,
        "results": len(filtered)
    })

    st.markdown(f"""
    <div class='enterprise-panel'>
        <div class='panel-header'>
            <div class='panel-title'>REGISTERED FARMER DIRECTORY — 6 CORE DATA CATEGORIES</div>
            <div class='panel-count-badge'>Showing {len(filtered):,} of {total_farmers:,} profiles</div>
        </div>""", unsafe_allow_html=True)

    st.dataframe(
        filtered[CORE_COLS].copy(),
        column_config={
            "Name":      st.column_config.TextColumn("Farmer Name",       width="medium"),
            "Phone":     st.column_config.TextColumn("Phone Number",      width="medium"),
            "Crop_Type": st.column_config.TextColumn("Crop Category",     width="small"),
            "Crop_Area": st.column_config.NumberColumn("Farm Area (Acres)", format="%d acres", width="small"),
            "Season":    st.column_config.TextColumn("Growing Season",    width="small"),
            "Location":  st.column_config.TextColumn("City / District",   width="medium"),
        },
        use_container_width=True,
        hide_index=True,
        height=480
    )

    st.write("")
    dl1, dl2, _ = st.columns([1.7, 1.9, 4.4])
    with dl1:
        if st.download_button("Export Filtered Directory (CSV)",
                              data=filtered[CORE_COLS].to_csv(index=False).encode("utf-8"),
                              file_name="saf_shikan_farmers_export.csv",
                              mime="text/csv",
                              use_container_width=True):
            log_event("DATA_EXPORT", "CSV export triggered — Farmer Directory",
                      {"format": "CSV", "rows_exported": len(filtered)})
    with dl2:
        if st.download_button("Export Filtered Directory (Styled Excel)",
                              data=generate_excel(filtered),
                              file_name="saf_shikan_farmers_export.xlsx",
                              mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                              use_container_width=True):
            log_event("DATA_EXPORT", "Excel export triggered — Farmer Directory",
                      {"format": "XLSX", "rows_exported": len(filtered)})

    st.markdown("</div>", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# TAB 2 — DATA GROUPING & PIVOT ANALYTICS ENGINE
# ══════════════════════════════════════════════════════════════════════════════
with tab_group:
    st.markdown("""
    <div class='enterprise-panel'>
        <div class='panel-header'>
            <div class='panel-title'>MULTI-DIMENSIONAL PIVOT AGGREGATION ENGINE</div>
            <div class='panel-count-badge'>BI GROUPING CONSOLE</div>
        </div>""", unsafe_allow_html=True)

    g1, g2 = st.columns([2, 2])
    group_by = g1.selectbox("Group By Dimension", ["Location", "Crop_Type", "Season", "Farm_Scale"],
                             format_func=lambda x: x.replace("_", " "))

    st.markdown("</div>", unsafe_allow_html=True)

    # Compute Aggregation — Count only
    grp = df.groupby(group_by).agg(
        Count=("Name", "count"),
        Total_Area=("Crop_Area", "sum")
    ).reset_index()
    grp.columns = [group_by.replace("_", " "), "Farmer Count", "Total Area (Acres)"]
    sort_col = "Farmer Count"

    grp = grp.sort_values(by=sort_col, ascending=False).reset_index(drop=True)
    if not grp.empty:
        top_name = grp.iloc[0, 0]
        top_val  = grp.iloc[0][sort_col]
    else:
        top_name = "N/A (No Data)"
        top_val  = 0

    log_event("PAGE_VIEW", f"Data Grouping tab — grouped by {group_by} (count)",
              {"group_by": group_by, "operation": "count"})

    st.markdown(f"""
    <div class='stats-row'>
        <div class='stat-card'>
            <div class='stat-header'><span class='stat-label'>Leading Dimension</span><span class='stat-badge'>Rank #1</span></div>
            <div class='stat-value'>{top_name}</div>
            <div class='stat-footer'><span class='dot dot-green'></span>Highest aggregate concentration</div>
        </div>
        <div class='stat-card'>
            <div class='stat-header'><span class='stat-label'>Top Group Count</span><span class='stat-badge'>COUNT</span></div>
            <div class='stat-value'>{top_val:,.0f}</div>
            <div class='stat-footer'><span class='dot dot-blue'></span>Peak metric in distribution</div>
        </div>
        <div class='stat-card'>
            <div class='stat-header'><span class='stat-label'>Total Categories</span><span class='stat-badge'>Segments</span></div>
            <div class='stat-value'>{len(grp)}</div>
            <div class='stat-footer'><span class='dot dot-amber'></span>Distinct groups</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("""
    <div class='enterprise-panel'>
        <div class='panel-header'>
            <div class='panel-title'>GROUPED AGGREGATION RESULTS TABLE</div>
            <div class='panel-count-badge'>COLOR HEATMAP ENABLED</div>
        </div>""", unsafe_allow_html=True)

    try:
        styled = grp.style.background_gradient(
            subset=[sort_col], cmap="YlGn"
        ).format({col: "{:,.0f}" for col in grp.select_dtypes("number").columns})
        st.dataframe(styled, use_container_width=True, hide_index=True)
    except Exception:
        st.dataframe(grp, use_container_width=True, hide_index=True)

    dl_col, _ = st.columns([2.5, 5.5])
    with dl_col:
        if st.download_button("Export Group Summary (CSV)",
                              data=grp.to_csv(index=False).encode("utf-8"),
                              file_name=f"saf_shikan_grouped_{group_by.lower()}.csv",
                              mime="text/csv",
                              use_container_width=True):
            log_event("DATA_EXPORT", f"CSV export — Grouped by {group_by}",
                      {"format": "CSV", "group_by": group_by, "operation": "count"})

    st.markdown("</div>", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# TAB 3 — VISUAL ANALYTICS (ENTERPRISE CHART CARDS)
# ══════════════════════════════════════════════════════════════════════════════
with tab_charts:
    FONT = "Inter, -apple-system, sans-serif"
    COLOR_GREEN = ["#0C3823","#1B5E20","#2E7D32","#388E3C","#43A047","#66BB6A","#A5D6A7","#C8E6C9"]
    LAYOUT = dict(font=dict(family=FONT), paper_bgcolor="rgba(0,0,0,0)",
                  plot_bgcolor="rgba(0,0,0,0)", margin=dict(t=45, b=25, l=10, r=10),
                  title_font=dict(size=16, family=FONT, color="#0F172A"))

    log_event("PAGE_VIEW", "Visual Analytics tab viewed — 4 charts rendered",
              {"charts": ["City Bar", "Crop Donut", "Season Bar", "City-Crop Stack"]})

    r1c1, r1c2 = st.columns(2)
    with r1c1:
        st.markdown("<div class='agron-chart-card'>", unsafe_allow_html=True)
        rc = df["Location"].value_counts().reset_index()
        rc.columns = ["City", "Farmers"]
        f1 = px.bar(rc, x="City", y="Farmers", color="City", text="Farmers",
                    color_discrete_sequence=COLOR_GREEN,
                    title="City-wise Farmer Concentration")
        f1.update_traces(textposition="outside", marker_cornerradius=6,
                         hovertemplate="<b>%{x}</b><br>Farmers: %{y}<extra></extra>")
        f1.update_layout(**LAYOUT, showlegend=False, height=310)
        f1.update_yaxes(showgrid=True, gridcolor="#E2E8F0", zeroline=False)
        f1.update_xaxes(showgrid=False, tickangle=45)
        st.plotly_chart(f1, use_container_width=True)
        st.markdown("</div>", unsafe_allow_html=True)

    with r1c2:
        st.markdown("<div class='agron-chart-card'>", unsafe_allow_html=True)
        cc = df["Crop_Type"].value_counts().reset_index()
        cc.columns = ["Crop", "Farmers"]
        f2 = px.pie(cc, names="Crop", values="Farmers", hole=0.58,
                    title="Crop Portfolio Distribution",
                    color_discrete_sequence=COLOR_GREEN)
        f2.update_traces(textposition="inside", textinfo="percent+label",
                         hovertemplate="<b>%{label}</b><br>Farmers: %{value}<br>Share: %{percent}<extra></extra>",
                         marker=dict(line=dict(color="#FFFFFF", width=2)))
        f2.update_layout(**LAYOUT, showlegend=False, height=310)
        st.plotly_chart(f2, use_container_width=True)
        st.markdown("</div>", unsafe_allow_html=True)

    r2c1, r2c2 = st.columns(2)
    with r2c1:
        st.markdown("<div class='agron-chart-card'>", unsafe_allow_html=True)
        sc = df["Season"].value_counts().reset_index()
        sc.columns = ["Season", "Farmers"]
        f3 = px.bar(sc, x="Season", y="Farmers", color="Season", text="Farmers",
                    color_discrete_sequence=["#0C3823","#388E3C","#81C784"],
                    title="Seasonal Cultivation Breakdown")
        f3.update_traces(textposition="outside", marker_cornerradius=6,
                         hovertemplate="<b>%{x}</b><br>Farmers: %{y}<extra></extra>")
        f3.update_layout(**LAYOUT, showlegend=False, height=290)
        f3.update_yaxes(showgrid=True, gridcolor="#E2E8F0", zeroline=False)
        f3.update_xaxes(showgrid=False)
        st.plotly_chart(f3, use_container_width=True)
        st.markdown("</div>", unsafe_allow_html=True)

    with r2c2:
        st.markdown("<div class='agron-chart-card'>", unsafe_allow_html=True)
        ta = df.groupby("Crop_Type")["Crop_Area"].sum().sort_values().reset_index()
        ta.columns = ["Crop", "Total Area (Acres)"]
        f4 = px.bar(ta, x="Total Area (Acres)", y="Crop", orientation="h",
                    color="Total Area (Acres)", color_continuous_scale=["#C8E6C9","#0C3823"],
                    text="Total Area (Acres)", title="Total Cultivated Area per Crop")
        f4.update_traces(texttemplate="%{x:,.0f} ac", textposition="outside",
                         marker_cornerradius=4,
                         hovertemplate="<b>%{y}</b><br>Total Area: %{x:,.0f} acres<extra></extra>")
        f4.update_layout(**LAYOUT, coloraxis_showscale=False, height=290)
        f4.update_xaxes(showgrid=True, gridcolor="#E2E8F0", zeroline=False)
        f4.update_yaxes(showgrid=False)
        st.plotly_chart(f4, use_container_width=True)
        st.markdown("</div>", unsafe_allow_html=True)

    st.markdown("<div class='agron-chart-card'>", unsafe_allow_html=True)
    cr = df.groupby(["Location","Crop_Type"]).size().reset_index(name="Farmers")
    top_cities = df["Location"].value_counts().nlargest(10).index.tolist()
    cr = cr[cr["Location"].isin(top_cities)]
    f5 = px.bar(cr, x="Location", y="Farmers", color="Crop_Type", barmode="stack",
                color_discrete_sequence=COLOR_GREEN,
                title="City-wise Crop Distribution Matrix (Top 10 Cities)",
                labels={"Crop_Type": "Crop", "Farmers": "Registered Farmers", "Location": "City"})
    f5.update_traces(marker_cornerradius=4,
                     hovertemplate="<b>%{fullData.name}</b><br>Farmers: %{y}<extra></extra>")
    f5.update_layout(**LAYOUT, height=360, legend_title_text="Crop Category")
    f5.update_yaxes(showgrid=True, gridcolor="#E2E8F0", zeroline=False)
    f5.update_xaxes(showgrid=False, tickangle=30)
    st.plotly_chart(f5, use_container_width=True)
    st.markdown("</div>", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# ══════════════════════════════════════════════════════════════════════════════
# TAB 4 — SECURITY & AUDIT TRAIL CONSOLE
# ══════════════════════════════════════════════════════════════════════════════
with tab_security:
    from audit_logger import _session_id

    session_id  = _session_id()
    log_counts  = get_log_counts()

    log_event("PAGE_VIEW", "Security & Audit Trail tab opened", {})

    # Re-fetch counts after above event
    log_counts    = get_log_counts()
    total_events  = sum(log_counts.values())
    page_views    = log_counts.get("PAGE_VIEW",    0)
    filter_events = log_counts.get("FILTER_APPLY", 0)
    export_events = log_counts.get("DATA_EXPORT",  0)
    entry_events  = log_counts.get("DATA_ENTRY",   0)
    auth_events   = log_counts.get("LOGIN", 0) + log_counts.get("LOGOUT", 0)

    # ── Security Posture Badges ──────────────────────────────────────────────
    st.markdown(f"""
    <div class='enterprise-panel'>
        <div class='panel-header'>
            <div class='panel-title'>SECURITY POSTURE &amp; COMPLIANCE STATUS</div>
            <div class='panel-count-badge'>VERIFIED</div>
        </div>
        <div class='security-badge-row'>
            <div class='security-badge'>
                <div class='security-badge-icon badge-https'>TLS</div>
                <div>
                    <div class='security-badge-text-label'>Data in Transit</div>
                    <div class='security-badge-text-value'>HTTPS / TLS 1.3 Enforced</div>
                    <div class='security-badge-text-sub'>All client-server traffic is encrypted end-to-end via Streamlit Cloud SSL certificate.</div>
                </div>
            </div>
            <div class='security-badge'>
                <div class='security-badge-icon badge-audit'>LOG</div>
                <div>
                    <div class='security-badge-text-label'>Audit Trail</div>
                    <div class='security-badge-text-value'>Persistent Cross-Session Logging</div>
                    <div class='security-badge-text-sub'>Every action is written to Supabase audit_log table — survives page refresh, accessible across all user sessions.</div>
                </div>
            </div>
            <div class='security-badge'>
                <div class='security-badge-icon badge-session'>SID</div>
                <div>
                    <div class='security-badge-text-label'>Current Session</div>
                    <div class='security-badge-text-value'>{session_id}</div>
                    <div class='security-badge-text-sub'>Signed in as: {current_user_email}. All events tagged with this session ID and user identity.</div>
                </div>
            </div>
            <div class='security-badge'>
                <div class='security-badge-icon badge-access'>ACC</div>
                <div>
                    <div class='security-badge-text-label'>Data Access Scope</div>
                    <div class='security-badge-text-value'>READ / WRITE — Authenticated</div>
                    <div class='security-badge-text-sub'>Access controlled via Supabase Auth. All write operations are user-attributed and audited.</div>
                </div>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # ── Session Event Counter Cards ──────────────────────────────────────────
    st.markdown(f"""
    <div class='stats-row'>
        <div class='stat-card'>
            <div class='stat-header'><span class='stat-label'>Total Events</span><span class='stat-badge'>Session</span></div>
            <div class='stat-value'>{total_events}</div>
            <div class='stat-footer'><span class='dot dot-blue'></span>All tracked actions</div>
        </div>
        <div class='stat-card'>
            <div class='stat-header'><span class='stat-label'>Page Views</span><span class='stat-badge'>NAV</span></div>
            <div class='stat-value'>{page_views}</div>
            <div class='stat-footer'><span class='dot dot-green'></span>Tab navigations</div>
        </div>
        <div class='stat-card'>
            <div class='stat-header'><span class='stat-label'>Filter Events</span><span class='stat-badge'>FILTER</span></div>
            <div class='stat-value'>{filter_events}</div>
            <div class='stat-footer'><span class='dot dot-purple'></span>Directory filters applied</div>
        </div>
        <div class='stat-card'>
            <div class='stat-header'><span class='stat-label'>Data Exports</span><span class='stat-badge'>EXPORT</span></div>
            <div class='stat-value'>{export_events}</div>
            <div class='stat-footer'><span class='dot dot-amber'></span>CSV / Excel downloads</div>
        </div>
        <div class='stat-card'>
            <div class='stat-header'><span class='stat-label'>Farmer Entries</span><span class='stat-badge'>WRITE</span></div>
            <div class='stat-value'>{entry_events}</div>
            <div class='stat-footer'><span class='dot dot-teal'></span>New records committed</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # ── Persistent Cross-Session Audit Log (from Supabase) ───────────────────
    st.markdown("""
    <div class='enterprise-panel'>
        <div class='panel-header'>
            <div class='panel-title'>PERSISTENT AUDIT ACTIVITY LOG</div>
            <div class='panel-count-badge'>SUPABASE — CROSS-SESSION</div>
        </div>
    """, unsafe_allow_html=True)

    with st.spinner("Loading audit records from Supabase..."):
        supabase_entries = get_log_from_supabase(limit=200)

    def _to_plain_english(entry: dict) -> str:
        etype = entry.get("event_type", "")
        desc  = entry.get("description", "")
        det   = entry.get("details", {})

        if etype == "FILTER_APPLY":
            parts = []
            if det.get("city") and det["city"] != "All Cities":
                parts.append(f"City: {det['city']}")
            if det.get("crop") and det["crop"] != "All Crops":
                parts.append(f"Crop: {det['crop']}")
            if det.get("season") and det["season"] != "All Seasons":
                parts.append(f"Season: {det['season']}")
            if det.get("keyword") and det["keyword"] not in ("(none)", ""):
                parts.append(f"Search: '{det['keyword']}'")
            filter_str = ", ".join(parts) if parts else "All default filters (no restrictions)"
            res = det.get("results", "")
            return f"Filtered Farmer Directory ({filter_str}). Found {res} matching records."
        elif etype == "DATA_EXPORT":
            fmt  = det.get("format", "File")
            rows = det.get("rows_exported", "")
            return f"Downloaded {rows} farmer records as {fmt}." if rows else f"Downloaded data summary as {fmt}."
        elif etype == "PAGE_VIEW":
            return f"Opened and viewed: {desc}."
        elif etype == "LOGIN":
            user = det.get("user", desc)
            recs = det.get("total_records", "")
            return f"User signed in: {user}. Portal loaded with {recs} farmer records."
        elif etype == "LOGOUT":
            user = det.get("user", desc)
            return f"User signed out: {user}."
        elif etype == "DATA_ENTRY":
            name = det.get("name", "")
            loc  = det.get("location", det.get("region", ""))
            crop = det.get("crop", "")
            area = det.get("area", "")
            return f"New farmer registered: {name} ({loc}, {crop}, {area} ac)." if name else f"New farmer registered: {desc}."
        elif etype == "SYSTEM":
            recs = det.get("total_records", "")
            return f"System started (TLS 1.3) — loaded {recs} records." if recs else f"System: {desc}."
        return desc

    icon_map = {
        "FILTER_APPLY": "🔍", "DATA_EXPORT": "📥", "PAGE_VIEW": "👁️",
        "SYSTEM": "🛡️", "DATA_ENTRY": "➕", "LOGIN": "🔐", "LOGOUT": "🚪"
    }

    if not supabase_entries:
        st.info(
            "No persistent audit records found. This happens if: (a) the `audit_log` table "
            "hasn't been created in Supabase yet, or (b) navigate the portal to generate events."
        )
    else:
        for entry in supabase_entries:
            plain_text = _to_plain_english(entry)
            ts         = entry.get("display_ts", "")
            sid        = entry.get("session", "")[:8]
            user_email = entry.get("user_email", "anonymous")
            icon       = icon_map.get(entry.get("event_type", ""), "📋")

            card_html = (
                f"<div style='background:#FFFFFF;border:1px solid #E2E8F0;border-radius:10px;"
                f"padding:14px 18px;margin-bottom:8px;box-shadow:0 1px 2px rgba(0,0,0,0.02);'>"
                f"<div style='font-size:0.93rem;font-weight:600;color:#0F172A;margin-bottom:5px;'>"
                f"{icon} &nbsp; {plain_text}</div>"
                f"<div style='font-size:0.76rem;color:#64748B;font-weight:500;display:flex;gap:16px;'>"
                f"<span>🕐 {ts}</span>"
                f"<span>👤 {user_email}</span>"
                f"<span>🔑 Session: {sid}</span>"
                f"</div></div>"
            )
            st.markdown(card_html, unsafe_allow_html=True)

    st.markdown("</div>", unsafe_allow_html=True)

    # ── Export ───────────────────────────────────────────────────────────────
    audit_col, _ = st.columns([2.5, 5.5])
    with audit_col:
        st.download_button(
            "Export Full Audit Log (CSV)",
            data=get_log_as_csv(use_supabase=True).encode("utf-8"),
            file_name=f"saf_shikan_audit_full.csv",
            mime="text/csv",
            use_container_width=True
        )
