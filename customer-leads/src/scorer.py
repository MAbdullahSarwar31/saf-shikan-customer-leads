"""Lead Scoring Module for SAF SHIKAN.

This module scores every customer, categorizes them as Hot/Warm/Cold/Low Priority,
and exports the final prioritized list to CSV and a formatted Excel sheet.
"""

import os
import sys
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Add parent directory to path so config can be imported when run directly
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from config import (
    CLUSTERED_DATA_PATH, FINAL_DATA_PATH, EXCEL_OUTPUT_PATH, LEAD_SCORING_WEIGHTS
)
from utils import load_dataset, save_dataset, logger, ensure_dirs


def calculate_lead_scores(df: pd.DataFrame) -> pd.DataFrame:
    """Calculate lead scores for all customers based on engineered feature scores.
    
    Formula:
      score = (
        Drone_Service_Potential * 0.35 +
        Area_Score * 0.30 +
        Season_Score * 0.20 +
        Region_Score * 0.15
      ) * 10
    
    Args:
        df: Input DataFrame containing feature scores.
        
    Returns:
        DataFrame with Lead_Score added.
    """
    logger.info("Calculating lead scores...")
    w = LEAD_SCORING_WEIGHTS
    
    # Formula execution
    df["Lead_Score"] = (
        df["Drone_Service_Potential"] * w["drone_potential"] +
        df["Area_Score"] * w["area"] +
        df["Season_Score"] * w["season"] +
        df["Region_Score"] * w["region"]
    ) * 10.0
    
    df["Lead_Score"] = df["Lead_Score"].round(2)
    return df


def categorize_leads(df: pd.DataFrame) -> pd.DataFrame:
    """Categorize leads based on scores and assign recommended actions.
    
    Categories:
      score >= 75 -> HOT
      score 50-74 -> WARM
      score 25-49 -> COLD
      score < 25  -> LOW PRIORITY
      
    Args:
        df: DataFrame with Lead_Score.
        
    Returns:
        DataFrame with Lead_Category, Priority_Rank, and Recommended_Action.
    """
    logger.info("Categorizing and ranking leads...")
    
    # 1. Categorize
    conditions = [
        (df["Lead_Score"] >= 75.0),
        (df["Lead_Score"] >= 50.0) & (df["Lead_Score"] < 75.0),
        (df["Lead_Score"] >= 25.0) & (df["Lead_Score"] < 50.0),
        (df["Lead_Score"] < 25.0)
    ]
    
    categories = ["HOT", "WARM", "COLD", "LOW PRIORITY"]
    df["Lead_Category"] = np.select(conditions, categories, default="COLD")
    
    # 2. Recommended Action
    actions = {
        "HOT": "Call within 48 hours — high drone service need",
        "WARM": "Schedule call this month — good potential",
        "COLD": "Add to WhatsApp broadcast list",
        "LOW PRIORITY": "Add to monthly newsletter only"
    }
    df["Recommended_Action"] = df["Lead_Category"].map(actions)
    
    # 3. Rank leads (descending by score)
    df["Priority_Rank"] = df["Lead_Score"].rank(ascending=False, method="min").astype(int)
    
    # Sort by Priority_Rank so best leads are at the top
    df = df.sort_values(by="Priority_Rank").reset_index(drop=True)
    return df


def export_to_excel_styled(df: pd.DataFrame, filepath: str) -> None:
    """Export scored leads to a beautifully styled Excel workbook using openpyxl.
    
    Args:
        df: Scored and categorized leads DataFrame.
        filepath: Output Excel path.
    """
    logger.info(f"Exporting styled Excel report to: {filepath}")
    
    # Selected columns for the sales team
    export_cols = [
        "Priority_Rank", "Name", "Phone", "Crop_Type", "Crop_Area", 
        "Season", "Location", "Farm_Scale", "Lead_Score", 
        "Lead_Category", "Recommended_Action"
    ]
    export_df = df[export_cols].copy()
    
    # Convert crop types back to title case for the Excel display
    export_df["Crop_Type"] = export_df["Crop_Type"].str.title()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Leads"
    
    # Enable grid lines
    ws.views.sheetView[0].showGridLines = True
    
    # Write Headers
    headers = [col.replace("_", " ") for col in export_cols]
    ws.append(headers)
    
    # Design elements
    # SAF SHIKAN Dark Green Brand Palette
    header_fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    # Category Fills and Fonts
    category_styles = {
        "HOT": {
            "fill": PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"), # Light Red
            "font": Font(name="Calibri", size=10, bold=True, color="B71C1C")                 # Dark Red
        },
        "WARM": {
            "fill": PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid"), # Light Orange
            "font": Font(name="Calibri", size=10, bold=True, color="E65100")                 # Dark Orange
        },
        "COLD": {
            "fill": PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"), # Light Blue
            "font": Font(name="Calibri", size=10, bold=True, color="0D47A1")                 # Dark Blue
        },
        "LOW PRIORITY": {
            "fill": PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"), # Light Gray
            "font": Font(name="Calibri", size=10, color="757575")                            # Medium Gray
        }
    }
    
    thin_border = Border(
        left=Side(style='thin', color='E0E0E0'),
        right=Side(style='thin', color='E0E0E0'),
        top=Side(style='thin', color='E0E0E0'),
        bottom=Side(style='thin', color='E0E0E0')
    )
    
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    
    # Format Headers
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
    
    # Append Data and apply styles row-by-row
    for _, row in export_df.iterrows():
        ws.append(list(row))
        curr_row = ws.max_row
        category = row["Lead_Category"]
        
        # Style all cells in the row
        for col_idx in range(1, len(export_cols) + 1):
            cell = ws.cell(row=curr_row, column=col_idx)
            cell.border = thin_border
            
            # Default Font & Alignment
            cell.font = Font(name="Calibri", size=10, color="000000")
            
            # Alignment rules
            if export_cols[col_idx-1] in ["Priority_Rank", "Phone", "Crop_Area", "Season", "Farm_Scale", "Lead_Score", "Lead_Category"]:
                cell.alignment = center_align
            else:
                cell.alignment = left_align
                
            # Formatting specifics
            if export_cols[col_idx-1] == "Lead_Score":
                cell.number_format = '0.00'
                
            # Apply category-specific colors to the Lead_Category column only (for elegant clean UI)
            if export_cols[col_idx-1] == "Lead_Category" and category in category_styles:
                cell.fill = category_styles[category]["fill"]
                cell.font = category_styles[category]["font"]
                
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        # Exclude headers from width calculation for columns that have long headers but short data
        for cell in col:
            val = str(cell.value or '')
            if cell.row > 1 and len(val) > max_len:
                max_len = len(val)
        
        # Factor in the header length with padding
        header_len = len(str(col[0].value or ''))
        width = max(max_len + 3, header_len + 4)
        ws.column_dimensions[col_letter].width = min(width, 40) # cap column width at 40
        
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    wb.save(filepath)
    logger.info(f"Excel report saved successfully at: {filepath}")


def main() -> None:
    """Main execution function to load clustered data, score, categorize, and export."""
    logger.info("Initializing Phase 4: Lead Scoring...")
    ensure_dirs()
    
    # Load clustered data
    df = load_dataset(CLUSTERED_DATA_PATH)
    
    # Score leads
    df_scored = calculate_lead_scores(df)
    
    # Categorize and rank leads
    df_final = categorize_leads(df_scored)
    
    # Save final dataset to CSV
    save_dataset(df_final, FINAL_DATA_PATH)
    
    # Save final dataset to styled Excel
    export_to_excel_styled(df_final, EXCEL_OUTPUT_PATH)
    
    # Print confirmations for verification gate
    print("=" * 60)
    print("PHASE 4 VERIFICATION GATE - LEAD SCORING SUMMARY")
    print("=" * 60)
    print(f"Total customers scored: {len(df_final)}")
    print(f"Excel file created: {os.path.exists(EXCEL_OUTPUT_PATH)} ({EXCEL_OUTPUT_PATH})")
    
    print("\nLead Score Distribution Details:")
    print(df_final["Lead_Score"].describe().loc[["min", "max", "mean", "std"]])
    
    print("\nLead Category Counts:")
    print(df_final["Lead_Category"].value_counts())
    
    print("\nTop 10 Hot Leads (High Drone Needs):")
    top_cols = ["Priority_Rank", "Name", "Phone", "Crop_Type", "Crop_Area", "Region", "Lead_Score", "Lead_Category"]
    print(df_final[top_cols].head(10).to_string(index=False))
    print("=" * 60)
    logger.info("Phase 4 Lead Scoring completed successfully.")


if __name__ == "__main__":
    main()
